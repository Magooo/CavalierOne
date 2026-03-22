import openpyxl
import sys
import os

def extract_excel_data(file_paths):
    all_text = ""
    for file_path in file_paths:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            all_text += f"\n--- File: {os.path.basename(file_path)} ---\n"
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                all_text += f"\n[Sheet: {sheet_name}]\n"
                for row in sheet.iter_rows(values_only=True):
                    # Filter out purely empty rows
                    if any(row):
                         # Convert None to empty string and join with tabs or commas
                         row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                         all_text += row_text + "\n"
        except Exception as e:
            all_text += f"Error reading {file_path}: {e}\n"
    
    return all_text

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_excel_content.py <file1> <file2> ...")
        sys.exit(1)
    
    files = sys.argv[1:]
    
    # Check if files exist
    valid_files = [f for f in files if os.path.exists(f)]
    if not valid_files:
        print("No valid files found.")
        sys.exit(1)

    extracted_content = extract_excel_data(valid_files)
    
    output_file = "extracted_excel.txt"
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(extracted_content)
    
    print(f"Extraction complete. Output written to {output_file}")
