import openpyxl
import os

def process_sheet(sheet, sheet_name):
    lines = []
    lines.append(f"# Sheet: {sheet_name}")
    
    # Identify header row? usually row 1 or 2. Let's just dump non-empty rows.
    for row in sheet.iter_rows(values_only=True):
        # Filter empty rows
        if not any(row):
            continue
        
        # Convert to string and clean
        row_str = [str(cell) if cell is not None else "" for cell in row]
        
        # Simple markdown table row if it looks like data
        # Or just pipe separated
        lines.append("| " + " | ".join(row_str) + " |")
    
    lines.append("\n")
    return "\n".join(lines)

def summarize_file(filepath, label):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    summary = []
    summary.append(f"# File: {label}")
    
    sheets_to_process = ["Fast Key", "Shepparton", "Echuca", "Investor", "Landscaping", "Value Packs", "Sandown 32 Display"]
    # Add NSW specific if names differ? Assuming similar structure or processing all relevant.
    
    for sheet_name in wb.sheetnames:
        # Filter irrelevant sheets
        if "Travel" in sheet_name or "Estimator" in sheet_name or "Working" in sheet_name:
            continue
        
        # If precise list is safer:
        # if sheet_name not in sheets_to_process: continue 
        # But names might vary slightly. Let's stick to exclusion list.
        
        summary.append(process_sheet(wb[sheet_name], sheet_name))
        
    return "\n".join(summary)

def main():
    base_dir = r"c:\Users\jason.m.chgv\Documents\CavalierOne\resources"
    output_file = os.path.join(base_dir, "fast_keys_summary.md")
    
    vic_file = os.path.join(base_dir, "Master - Fast Keys Sales Estimate - VIC.xlsx")
    nsw_file = os.path.join(base_dir, "Master - Fast Keys Sales Estimate - NSW.xlsx")
    
    full_content = ""
    
    if os.path.exists(vic_file):
        full_content += summarize_file(vic_file, "VIC Data") + "\n---\n"
        
    if os.path.exists(nsw_file):
        full_content += summarize_file(nsw_file, "NSW Data")
        
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(full_content)
    
    print(f"Summary written to {output_file}")

if __name__ == "__main__":
    main()
