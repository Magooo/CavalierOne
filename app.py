import os
import json
from datetime import timedelta
from flask import Flask, render_template, request, jsonify, send_from_directory, session, redirect, url_for, Response
from werkzeug.middleware.proxy_fix import ProxyFix
from dotenv import load_dotenv
import markdown
from functools import wraps

load_dotenv() # Load environment variables from .env

app = Flask(__name__)

# Vercel terminates SSL at the edge and forwards requests to Flask over HTTP.
# ProxyFix tells Flask to trust the X-Forwarded-Proto header so that
# request.url, url_for(), and session cookies all use https:// correctly.
# Without this, post-login redirects go to http:// and browsers drop cookies.
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# IMPORTANT: FLASK_SECRET must be set as a stable env var in Vercel.
# If missing, os.urandom(24) generates a NEW key on every cold-start,
# instantly invalidating all user sessions.
app.secret_key = os.environ.get('FLASK_SECRET', os.urandom(24))

# Keep users logged in for 30 days
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False  # Vercel handles HTTPS at the edge
app.config['SESSION_COOKIE_HTTPONLY'] = True
# Vercel serverless has a hard 4.5 MB request body limit.
# Locally (no VERCEL env var) we allow up to 50 MB — full-quality renders are fine.
IS_VERCEL = bool(os.environ.get('VERCEL'))
UPLOAD_LIMIT_MB = 4 if IS_VERCEL else 50
app.config['MAX_CONTENT_LENGTH'] = UPLOAD_LIMIT_MB * 1024 * 1024

@app.errorhandler(413)
def request_entity_too_large(e):
    return jsonify({
        'error': f'File too large. Maximum upload size is {UPLOAD_LIMIT_MB} MB on this server. '
                 f'{"Please resize your image (e.g. save as JPEG at 80% quality) before uploading." if IS_VERCEL else "Try a different format or reduce resolution."}'
    }), 413

@app.errorhandler(500)
def internal_error(e):
    """Temporary: expose traceback so we can diagnose Vercel startup crashes."""
    import traceback
    tb = traceback.format_exc()
    return f"<pre style='padding:2rem;font-family:monospace;white-space:pre-wrap'><b>500 Internal Server Error</b>\n\n{tb}</pre>", 500

@app.route('/api/health')
def health_check():
    """Diagnostic endpoint — shows startup import errors and Python version."""
    import sys
    errors = globals().get('_STARTUP_ERRORS', [])
    return jsonify({
        'status': 'ok' if not errors else 'degraded',
        'python': sys.version,
        'startup_errors': errors,
        'is_vercel': bool(os.environ.get('VERCEL')),
    })


import io

# --- Supabase Role-Based Auth Middleware ---
AUTH_ENABLED = os.environ.get('SUPABASE_AUTH_ENABLED', 'true').lower() == 'true'

# --- Permissions Mapping ---
ROLE_PERMISSIONS = {
    'admin': ['all'],
    'sales': ['house_land', 'sales_estimate', 'rendering'],
    'marketing': ['social_media', 'youtube_miner', 'doc_formatter', 'rendering', 'house_land']
}

def user_can_access(tool_id):
    role = session.get('role', 'guest')
    if role == 'admin': return True
    return tool_id in ROLE_PERMISSIONS.get(role, [])

@app.context_processor
def inject_permissions():
    """Makes user_can_access available in all jinja templates via Jinja."""
    return dict(user_can_access=user_can_access, session=session)

def require_role(roles):
    """Decorator to restrict access based on user role."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not AUTH_ENABLED:
                return f(*args, **kwargs)
            user_role = session.get('role', 'guest')
            if user_role == 'admin':
                return f(*args, **kwargs)
            if isinstance(roles, list) and user_role not in roles:
                # Return JSON for API routes so frontend can parse the error
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Access Denied: Insufficient permissions.'}), 403
                return "Access Denied: Insufficient permissions.", 403
            elif not isinstance(roles, list) and user_role != roles:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Access Denied: Insufficient permissions.'}), 403
                return "Access Denied: Insufficient permissions.", 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.before_request
def make_session_permanent():
    """Mark every session as permanent so the 30-day lifetime applies."""
    session.permanent = True

@app.before_request
def global_auth_check():
    if not AUTH_ENABLED:
        return
    # Exclude login, static assets, API routes, and the render endpoint
    # (/render_document just converts markdown POSTed in the body — no protected data)
    excluded = ('/login', '/static', '/resources', '/api/', '/render_document')
    if any(request.path.startswith(p) for p in excluded):
        return
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.url))
        
    # App-Level Multi-Tenancy Firewall
    # Blocks users who share the Supabase project but aren't assigned a CavalierOne role
    if session.get('role', 'guest') == 'guest':
        return Response(
            "Access Denied: You are authenticated, but you do not have an active role in CavalierOne. "
            "If you need access, ask your Admin to assign you a role in the CavalierOne Admin Panel.", 403
        )
# -----------------------------------------------------------------------------
# Safe module imports — each wrapped so one failure doesn't crash the whole app.
# Errors are captured in _STARTUP_ERRORS and exposed via /api/health.
# -----------------------------------------------------------------------------
_STARTUP_ERRORS = []

def _safe_import(label, fn):
    try:
        return fn()
    except Exception as _e:
        import traceback as _tb
        _STARTUP_ERRORS.append(f"[{label}] {_e}\n{_tb.format_exc()}")
        return None

# Config
_config_result = _safe_import("config", lambda: __import__('config').Config)
Config = _config_result

# prompt_builder
_pb = _safe_import("prompt_builder", lambda: __import__('utils.prompt_builder', fromlist=['build_marketing_prompt']))
build_marketing_prompt = getattr(_pb, 'build_marketing_prompt', None) if _pb else None

# youtube_miner
_ym = _safe_import("youtube_miner", lambda: __import__('utils.youtube_miner', fromlist=['get_channel_videos', 'get_video_transcript']))
get_channel_videos  = getattr(_ym, 'get_channel_videos',  None) if _ym else None
get_video_transcript = getattr(_ym, 'get_video_transcript', None) if _ym else None

# data_loader
_dl = _safe_import("data_loader", lambda: __import__('utils.data_loader', fromlist=['load_json_template', 'load_master_prompt', 'get_style_guide', 'get_company_info']))
load_json_template = getattr(_dl, 'load_json_template', None) if _dl else None
load_master_prompt  = getattr(_dl, 'load_master_prompt',  None) if _dl else None
get_style_guide     = getattr(_dl, 'get_style_guide',     None) if _dl else None
get_company_info    = getattr(_dl, 'get_company_info',    None) if _dl else None

# marketing
_mkt = _safe_import("marketing", lambda: __import__('marketing', fromlist=['SocialMediaPost']))
SocialMediaPost = getattr(_mkt, 'SocialMediaPost', None) if _mkt else None

# notebook_client / fireflies_client
_nc = _safe_import("notebook_client", lambda: __import__('utils.notebook_client', fromlist=['NotebookLMClient']))
NotebookLMClient = getattr(_nc, 'NotebookLMClient', None) if _nc else None

_fc = _safe_import("fireflies_client", lambda: __import__('utils.fireflies_client', fromlist=['FirefliesClient']))
FirefliesClient = getattr(_fc, 'FirefliesClient', None) if _fc else None

# Supabase
from supabase import create_client, Client
import tempfile
import uuid

supabase_url = os.environ.get("SUPABASE_URL")
supabase_key = os.environ.get("SUPABASE_KEY")
supabase: Client = None
if supabase_url and supabase_key:
    supabase = create_client(supabase_url, supabase_key)

# Initialize Clients (guard against failed imports)
notebook_client  = NotebookLMClient()  if NotebookLMClient  else None
fireflies_client = FirefliesClient()   if FirefliesClient   else None

@app.route('/login', methods=['GET', 'POST'])
def login():
    if not AUTH_ENABLED:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        try:
            res = supabase.auth.sign_in_with_password({"email": email, "password": password})
            session['user_id'] = res.user.id
            
            # Fetch user role
            try:
                roles = supabase.table('user_roles').select('*').eq('user_id', res.user.id).execute()
                session['role'] = roles.data[0]['role'] if roles.data else 'guest'
            except Exception as e:
                session['role'] = 'guest'
                print(f"Role fetch error: {e}")
                
            return redirect(request.args.get('next') or url_for('index'))
        except Exception as e:
            error_msg = str(e)
            if "Invalid login credentials" in error_msg:
                error_msg = "Invalid email or password."
            return render_template('login.html', error=error_msg)
            
    return render_template('login.html')

@app.route('/settings', methods=['GET', 'POST'])
def user_settings():
    if not AUTH_ENABLED:
        return redirect(url_for('index'))
        
    success = None
    if request.method == 'POST':
        env_text = request.form.get('env_text', '')
        if 'env_file' in request.files and request.files['env_file'].filename:
            env_text = request.files['env_file'].read().decode('utf-8')
        
        if env_text:
            import dotenv
            parsed_env = dotenv.dotenv_values(stream=io.StringIO(env_text))
            current_env = session.get('user_env', {})
            current_env.update(parsed_env)
            session['user_env'] = current_env
            session.modified = True
            success = "Environment configured successfully."
            
    return render_template('settings.html', user_env=session.get('user_env', {}), success=success)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/admin', methods=['GET', 'POST'])
@require_role('admin')
def admin_panel():
    error = None
    success = None
    if request.method == 'POST':
        action = request.form.get('action', 'update_role') # Default for backward compatibility
        
        if action == 'update_role':
            user_id = request.form.get('user_id')
            new_role = request.form.get('role')
            if user_id and new_role:
                try:
                    supabase.rpc('update_user_role', {"p_user_id": user_id, "p_role": new_role}).execute()
                    if user_id == session.get('user_id'):
                        session['role'] = new_role
                    success = "Role updated successfully."
                except Exception as e:
                    error = f"Update failed. Ensure you are an Admin. Details: {e}"
                    
        elif action == 'add_user':
            new_email = request.form.get('new_email').strip()
            new_role = request.form.get('new_role')
            if new_email and new_role:
                try:
                    supabase.rpc('grant_role_by_email', {"p_email": new_email, "p_role": new_role}).execute()
                    success = f"Granted {new_role} access to {new_email} successfully!"
                except Exception as e:
                    error_msg = str(e)
                    if 'User email not found' in error_msg:
                        error = "Email not found! They must sign in or be invited via Supabase first."
                    else:
                        error = f"Failed to add user: {e}"

    users = []
    try:
        # Fetch users + roles securely
        res = supabase.rpc('get_users_with_roles').execute()
        users = res.data
    except Exception as e:
        error = f"Fetch failed: {e}"

    return render_template('admin.html', users=users, error=error, success=success)

@app.route('/', methods=['GET', 'POST'])
def index():
    # Load library data
    home_designs = load_json_template('home_design_template')
    land_info = load_json_template('land_details_template')
    
    # Load Brand Config
    brand_config_path = os.path.join("resources", "brand_config.json")
    brand_config = {}
    if os.path.exists(brand_config_path):
        import json
        with open(brand_config_path, 'r') as f:
            brand_config = json.load(f)

    # Ensure they are lists for iteration (if template returns single dict, wrap it)
    # Ideally, input_templates.json should store lists. For now, we wrap single items.
    if isinstance(home_designs, dict): home_designs = [home_designs]
    if isinstance(land_info, dict): land_info = [land_info]
    
    if request.method == 'POST':
        form_type = request.form.get('form_type', 'house_land')
        
        # Enforce Backend RBAC
        if not user_can_access(form_type):
            return render_template('output.html', prompt=None, content=f"<div class='error-box'>Access Denied: You do not have permission to execute the {form_type} tool.</div>", data={'media_type': 'Error', 'home_name': 'Permission Denied'})
        
        final_prompt = ""
        user_data = {}
        content_list = [] # List to store multiple generated posts for YouTube
        
        if form_type == 'youtube_miner':
            # YOUTUBE MINER HANDLER
            channel_url = request.form.get('channel_url')
            count = int(request.form.get('video_count', 3))
            tone = request.form.get('yt_tone', 'Professional')
            
            # 1. Fetch Videos (or use selected)
            selected_video_id = request.form.get('selected_video_id')
            selected_video_title = request.form.get('selected_video_title')
            
            videos = []
            if selected_video_id:
                # User selected a specific video
                videos = [{'video_id': selected_video_id, 'title': selected_video_title or 'Selected Video', 'url': f"https://www.youtube.com/watch?v={selected_video_id}"}]
            else:
                # Fallback: Fetch latest N
                videos = get_channel_videos(channel_url, limit=count)
            
            print(f"Processing videos: {videos}")
            
            # 2. Process each video
            # Reuse class for generation
            # Or direct LLM call? Let's use SocialMediaPost for consistency but we need to inject transcript.
            # Actually, let's just make a new prompt here for simplicity or add a method to SocialMediaPost
            
            from marketing.llm_client import LLMClient
            client = LLMClient(provider="openai")
            
            for vid in videos:
                transcript = get_video_transcript(vid['video_id'])
                if not transcript:
                    continue # Skip if no transcript
                
                # Truncate transcript to avoid token limits (approx 10k chars)
                transcript_preview = transcript[:10000]
                
                prompt = f"""
---
## CONTEXT
**Source Material:** YouTube Video Transcript
**Video Title:** {vid['title']}
**Video URL:** {vid['url']}

## TRANSCRIPT (Excerpt)
{transcript_preview}

## INSTRUCTION
Write a high-quality Facebook/Instagram post to promote this video.
Tone: {tone}
1. Start with a hook based on the video content.
2. Summarize the key takeaway.
3. Include the video link: {vid['url']}
4. Add 3 hashtags.
"""
                
                generated_text = client.generate_text(prompt)
                
                # HTML Format the text
                html_text = markdown.markdown(generated_text)
                
                content_list.append({
                    'title': vid['title'],
                    'url': vid['url'],
                    'content': html_text,
                    'orig_prompt': prompt
                })
            
            # We need a special output view for lists
            return render_template('output_youtube.html', posts=content_list)

        elif form_type == 'social_media':
            # ... (existing social media logic) ...
            # SOCIAL MEDIA HANDLER
            platform = request.form.get('platform')
            tone = request.form.get('tone')
            raw_content = request.form.get('raw_content')
            
            # Handle File Upload
            image_file = request.files.get('post_image')
            image_path = None
            display_image = None
            if image_file and image_file.filename:
                # Upload directly to Supabase Storage
                if supabase:
                    file_ext = os.path.splitext(image_file.filename)[1]
                    unique_filename = f"{uuid.uuid4().hex}{file_ext}"
                    filepath = f"public/{unique_filename}"
                    
                    # Read the binary data from Flask FileStorage object
                    file_bytes = image_file.read()
                    
                    supabase.storage.from_("cavalierone_uploads").upload(
                        filepath, file_bytes, {"content-type": image_file.content_type}
                    )
                    
                    # Get public URL
                    public_url = supabase.storage.from_("cavalierone_uploads").get_public_url(filepath)
                    image_path = public_url
                    display_image = public_url
                    print(f"Image uploaded to Supabase: {public_url}")
                else:
                    print("Supabase not configured, skipping image upload.")
            
            post = SocialMediaPost(
                platform=platform,
                raw_content=raw_content,
                tone=tone,
                image_path=image_path
            )
            final_prompt = post.generate_content()
            
            # Convert Markdown to HTML for display
            html_content = markdown.markdown(final_prompt)
            
            user_data = {
                'media_type': f"Social Post ({platform})", 
                'tone': tone, 
                'notes': raw_content, 
                'image': display_image
            }
            # Use 'prompt' for the raw text (for copying) and 'content' for the visual preview
            return render_template('output.html', prompt=final_prompt, content=html_content, data=user_data)
        
        elif form_type == 'doc_formatter':
            title = request.form.get('doc_title', 'Generated Document')
            subtitle = request.form.get('doc_subtitle', '')
            style = request.form.get('doc_style', 'Standard Manual')
            raw_content = request.form.get('raw_doc_content', '')
            doc_context = request.form.get('doc_context', 'none')
            
            from marketing.llm_client import LLMClient
            client = LLMClient(provider="openai")
            
            context_string = ""
            if doc_context == 'fast_keys':
                try:
                    # Load fast_keys_summary.md as the full fast_keys_final.md is huge (400KB+)
                    with open(os.path.join('resources', 'fast_keys_summary.md'), 'r', encoding='utf-8') as f:
                        # If fast_keys_summary doesn't exist, fallback to reading a chunk of fast_keys_final
                        context_string = f.read()[:50000] 
                except FileNotFoundError:
                    try:
                        with open(os.path.join('resources', 'fast_keys_final.md'), 'r', encoding='utf-8') as f:
                            context_string = f.read()[:50000]
                    except Exception as e:
                        context_string = f"Error loading Fast Keys Final: {e}"
                except Exception as e:
                    context_string = f"Error loading Fast Keys: {e}"
            elif doc_context == 'master_prompt':
                try:
                    with open('C1_MASTER_PROMPT.md', 'r', encoding='utf-8') as f:
                        context_string = f.read()[:50000]
                except Exception as e:
                    context_string = f"Error loading Master Prompt: {e}"

            if style == 'Meeting Notes Processor':
                # Fireflies already produces a perfect summary — no AI reprocessing needed.
                # Just add the Cavalier One header and send straight to the edit page.
                meeting_attendees = request.form.get('meeting_attendees', '').strip()
                additional_instructions = request.form.get('meeting_extra_instructions', '').strip()

                header_lines = [
                    f"# {title}",
                    "",
                    "**Cavalier One | Meeting Notes**",
                    "",
                ]
                if meeting_attendees:
                    header_lines += ["## Attendees", meeting_attendees, ""]
                if additional_instructions:
                    header_lines += [f"> **Note:** {additional_instructions}", ""]

                header_lines.append("---")
                header_lines.append("")

                content_body = raw_content if raw_content else "_No meeting notes provided._"
                generated_text = "\n".join(header_lines) + "\n" + content_body

                return render_template(
                    'edit_draft.html',
                    title=title,
                    subtitle=subtitle,
                    raw_markdown=generated_text
                )

            else:
                prompt = f"""
---
## INSTRUCTION
You are an expert technical writer and document formatter for Cavalier Homes. 
Take the instructions and context below and generate a clean, structured Markdown document. 
Do not include any conversational filler like "Here is the document".
Output ONLY the formatted document. Let the user's "Extra Notes" guide WHAT you extract from the internal context. If no extra notes are provided, summarize the internal context in the style requested.

## FORMAT STYLE requested: {style}
- If 'Standard Manual', use Headings, bullet points, and clear paragraphs.
- If 'Data Table', organize the data logically into Markdown tables.
- If 'Step by Step Guide', use numbered lists and clear sections.
- If 'Sales Framework', use persuasive tone, actionable steps, and highlighting.

## INTERNAL CONTEXT PROVIDED
{context_string if context_string else "None provided."}

## EXTRA NOTES / INSTRUCTIONS (RAW DATA)
{raw_content if raw_content else "No extra instructions. Summarize the internal context above into a useful document based on the requested Format Style."}
"""

            generated_text = client.generate_text(prompt)

            # Send to EDIT page first — user edits markdown before final render
            return render_template(
                'edit_draft.html',
                title=title,
                subtitle=subtitle,
                raw_markdown=generated_text
            )
            
        else:
            # LEGACY / HOUSE & LAND HANDLER
            data = {
                'media_type': request.form.get('media_type'),
                'home_name': request.form.get('home_name'),
                'inclusions': request.form.get('inclusions'),
                'land_details': request.form.get('land_details'),
                'target_audience': request.form.get('target_audience'),
                'state': request.form.get('state'),
                'examples': request.form.get('examples')
            }
            
            if data['media_type'] == 'Social Media Content':
                # OLD LOGIC - Kept for compatibility if selected via dropdown in Tab 1
                post = SocialMediaPost(
                    platform="Generic",
                    home_name=data['home_name'],
                    target_audience=data['target_audience'],
                    land_details=data['land_details'],
                    inclusions=data['inclusions']
                )
                final_prompt = post.generate_content()
            else:
                final_prompt = build_marketing_prompt(data)
            
            user_data = data
        
            return render_template('output.html', prompt=final_prompt, content=None, data=user_data)
        
    return render_template('index.html', home_designs=home_designs, land_info=land_info, brand=brand_config)

import time
from flask import send_file
from utils.pdf_generator import create_brochure

@app.route('/api/get_channel_videos')
def get_channel_videos_api():
    channel_url = request.args.get('channel_url')
    limit = int(request.args.get('limit', 10))
    
    if not channel_url:
        return jsonify({'error': 'Channel URL is required'}), 400
        
    videos = get_channel_videos(channel_url, limit=limit)
    return jsonify(videos)


@app.route('/api/generate-image', methods=['POST'])
@require_role(['admin', 'marketing'])
def api_generate_image():
    """
    Generate a Cavalier-branded marketing image via kie.ai.

    POST body (JSON):
      image_type:  'listing_hero' | 'interior' | 'social_feed' | 'brochure'
      model:       (optional) kie.ai model slug. Defaults to flux-2-flex-text-to-image
      suburb:      (for listing_hero) e.g. 'Shepparton'
      house_type:  (for listing_hero) e.g. 'single storey'
      room:        (for interior) e.g. 'kitchen'
      format:      (for social_feed) 'square' | 'portrait' | 'landscape' | 'story'
      caption_context: (for social_feed) brief description of what the post is about
      section_title: (for brochure) document section name
      prompt_override: (optional) full custom prompt, bypasses template builder
      width:       (optional) image width in pixels, default 1024
      height:      (optional) image height in pixels, default 768

    Returns JSON:
      { image_url, image_urls, prompt, model, image_type }
    """
    import traceback
    from marketing.image_generator import generate_cavalier_image, build_listing_hero_prompt, \
        build_interior_lifestyle_prompt, build_social_feed_prompt, build_brochure_section_prompt
    from utils.kie_client import generate_image as kie_generate

    try:
        data = request.get_json(force=True, silent=True) or {}

        image_type = data.get('image_type', 'listing_hero')
        model = data.get('model', 'flux-2/flex-text-to-image')
        aspect_ratio = data.get('aspect_ratio', '16:9')
        resolution = data.get('resolution', '1K')
        prompt_override = data.get('prompt_override', '').strip()

        if prompt_override:
            # Use the custom prompt directly
            image_urls = kie_generate(
                prompt=prompt_override,
                model=model,
                aspect_ratio=aspect_ratio,
                resolution=resolution,
            )
            result = {
                'image_url': image_urls[0] if image_urls else None,
                'image_urls': image_urls,
                'prompt': prompt_override,
                'model': model,
                'image_type': image_type,
            }
        else:
            # Build kwargs for the appropriate prompt template
            kwargs = {}
            if image_type == 'listing_hero':
                kwargs = {
                    'suburb': data.get('suburb', ''),
                    'house_type': data.get('house_type', ''),
                    'extras': data.get('extras', ''),
                }
            elif image_type == 'interior':
                kwargs = {
                    'room': data.get('room', 'kitchen'),
                    'extras': data.get('extras', ''),
                }
            elif image_type == 'social_feed':
                kwargs = {
                    'caption_context': data.get('caption_context', ''),
                    'format': data.get('format', 'square'),
                    'extras': data.get('extras', ''),
                }
            elif image_type == 'brochure':
                kwargs = {
                    'section_title': data.get('section_title', ''),
                    'extras': data.get('extras', ''),
                }

            result = generate_cavalier_image(
                image_type=image_type,
                model=model,
                aspect_ratio=aspect_ratio,
                resolution=resolution,
                **kwargs
            )


        return jsonify(result)

    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        return jsonify({'error': f'Unexpected error: {str(e)}', 'detail': traceback.format_exc()[-600:]}), 500


@app.route('/api/upload-image', methods=['POST'])
@require_role(['admin', 'marketing'])
def api_upload_image():
    """
    Upload an image file to Supabase Storage and return the public URL.

    Used by the Image Studio to get a publicly-accessible URL before
    passing the image to KIE.ai for enhancement.

    Form field: file  (multipart/form-data)
    Returns JSON: { url, filename }
    """
    import traceback

    try:
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'error': 'No file provided'}), 400

        allowed = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
        ext = os.path.splitext(file.filename)[1].lstrip('.').lower()
        if ext not in allowed:
            return jsonify({'error': f'File type .{ext} not allowed. Use: {", ".join(allowed)}'}), 400

        file_bytes = file.read()
        unique_name = f"studio/{uuid.uuid4().hex}.{ext}"

        if supabase:
            supabase.storage.from_("cavalierone_uploads").upload(
                unique_name, file_bytes, {"content-type": file.content_type or f"image/{ext}"}
            )
            public_url = supabase.storage.from_("cavalierone_uploads").get_public_url(unique_name)
        else:
            # Fallback for local dev: save to disk and serve via Flask
            # KIE.ai requires a real https:// or http:// URL — data URIs are rejected
            upload_dir = os.path.join(os.getcwd(), 'resources', 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            local_filename = f"{uuid.uuid4().hex}.{ext}"
            local_path = os.path.join(upload_dir, local_filename)
            with open(local_path, 'wb') as f_out:
                f_out.write(file_bytes)
            # Build absolute URL so KIE.ai can fetch it
            public_url = request.host_url.rstrip('/') + f'/uploads/{local_filename}'
            unique_name = local_filename

        return jsonify({'url': public_url, 'filename': unique_name})

    except Exception as e:
        return jsonify({'error': f'Upload failed: {str(e)}', 'detail': traceback.format_exc()[-400:]}), 500


@app.route('/api/enhance-image', methods=['POST'])
@require_role(['admin', 'marketing'])
def api_enhance_image():
    """
    Enhance an uploaded photo using KIE.ai image-to-image.

    POST body (JSON):
      image_url:    Publicly accessible URL of the source image
      prompt:       What to turn the image into (e.g. "bright modern Australian home exterior")
      style:        Preset style key: 'cavalier_brand' | 'social_media' | 'brochure' | 'custom'
      strength:     0.0–1.0, how much to change (default 0.75)
      aspect_ratio: Output aspect ratio (default "16:9")
      model:        kie.ai img2img model slug (default "flux-2/flex-image-to-image")

    Returns JSON: { image_url, image_urls, prompt, model }
    """
    import traceback
    from utils.kie_client import enhance_image as kie_enhance

    # Style presets — inject brand-consistent phrasing into the prompt
    STYLE_PRESETS = {
        'cavalier_brand': (
            "modern Australian residential home, fresh bright natural lighting, "
            "lush green lawn, clean contemporary architecture, photorealistic, "
            "professional real estate photography, warm inviting atmosphere"
        ),
        'social_media': (
            "aspirational lifestyle photograph, vibrant colours, instagram-worthy, "
            "modern Australian home, natural sunlight, premium feel, no text or watermarks"
        ),
        'brochure': (
            "clean minimal brochure photograph, bright airy modern interior, "
            "white and neutral tones, wide letterbox crop, professional, photorealistic"
        ),
        'custom': '',
    }

    try:
        data = request.get_json(force=True, silent=True) or {}

        image_url = data.get('image_url', '').strip()
        if not image_url:
            return jsonify({'error': 'image_url is required'}), 400

        style = data.get('style', 'cavalier_brand')
        custom_prompt = data.get('prompt', '').strip()
        style_prefix = STYLE_PRESETS.get(style, STYLE_PRESETS['cavalier_brand'])

        # Merge style prefix with any user-specified prompt detail
        if custom_prompt and style_prefix:
            prompt = f"{style_prefix}, {custom_prompt}"
        elif custom_prompt:
            prompt = custom_prompt
        else:
            prompt = style_prefix

        model = data.get('model', 'flux-2/flex-image-to-image')
        strength = float(data.get('strength', 0.75))
        aspect_ratio = data.get('aspect_ratio', '16:9')

        image_urls = kie_enhance(
            image_url=image_url,
            prompt=prompt,
            model=model,
            strength=strength,
            aspect_ratio=aspect_ratio,
        )

        return jsonify({
            'image_url': image_urls[0] if image_urls else None,
            'image_urls': image_urls,
            'prompt': prompt,
            'model': model,
            'original_url': image_url,
        })

    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        return jsonify({'error': f'Unexpected error: {str(e)}', 'detail': traceback.format_exc()[-600:]}), 500


@app.route('/resources/brand_assets/<path:filename>')
def serve_brand_assets(filename):
    directory = os.path.join(os.getcwd(), 'resources', 'brand_assets')
    return send_from_directory(directory, filename)

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_file(os.path.join("resources", "uploads", filename))



@app.route('/api/generate-from-plans', methods=['POST'])
@require_role(['admin', 'marketing'])
def api_generate_from_plans():
    """
    Generate a photorealistic facade render from an architectural elevation drawing.

    Multipart form fields:
      elevation_image  — the elevation drawing file (PNG, JPG, WebP)
      style            — house style e.g. "Modern Australian Contemporary"
      notes            — free text notes from the user
      material_prompt  — swatch-derived material descriptions

    Returns JSON: { image_url, vision_analysis, prompt, engine }
    """
    import traceback

    try:
        file = request.files.get('elevation_image')
        if not file or not file.filename:
            return jsonify({'error': 'No elevation image uploaded'}), 400

        allowed_ext = {'png', 'jpg', 'jpeg', 'webp'}
        ext = os.path.splitext(file.filename)[1].lstrip('.').lower()
        if ext not in allowed_ext:
            return jsonify({'error': f'Unsupported file type .{ext}. Use PNG, JPG, or WebP.'}), 400

        style         = request.form.get('style', 'Modern Australian Contemporary')
        notes         = request.form.get('notes', '')
        material_prompt = request.form.get('material_prompt', '')
        # Painted zones: comma-separated list of zone labels the user painted on the elevation
        # e.g. "Roof,Brick,Render" — used to tell GPT-4 Vision which zones were colour-coded
        painted_zones_raw = request.form.get('painted_zones', '')
        painted_zones = [z.strip() for z in painted_zones_raw.split(',') if z.strip()]

        # Save to system temp dir — Vercel serverless only allows writes to /tmp
        # resources/uploads/ is in the read-only task bundle on Vercel (Errno 30)
        import tempfile
        upload_dir = tempfile.gettempdir()  # /tmp on Linux/Vercel, %TEMP% locally
        safe_name  = f"plan_{uuid.uuid4().hex}.{ext}"
        file_path  = os.path.join(upload_dir, safe_name)
        file.save(file_path)

        # ── Step 1: GPT-4 Vision reads the elevation ───────────────────────────
        from marketing.llm_client import LLMClient
        client = LLMClient(provider="openai")

        # Build zone colour key to help Vision read painted zones
        zone_colour_map = {
            'Roof':    'dark charcoal/grey zones',
            'Brick':   'red-brown/terracotta zones',
            'Render':  'warm cream/beige zones',
            'Feature': 'steel blue zones',
            'Windows': 'near-black zones',
            'Garage':  'sandy beige zones',
        }
        zone_context = ''
        if painted_zones:
            zone_lines = []
            for z in painted_zones:
                colour_hint = zone_colour_map.get(z, f'{z} zones')
                zone_lines.append(f"- {colour_hint} on the drawing = {z} surface area")
            zone_context = (
                "\n\nIMPORTANT — the user has colour-coded specific zones on this elevation drawing:\n"
                + "\n".join(zone_lines)
                + "\nUse these coloured zones to understand WHICH PART of the facade each material should be applied to."
            )

        vision_prompt = f"""You are a professional architectural renderer.
I will show you an architectural elevation drawing of a residential house. Your task is to produce an ARCHITECTURALLY PRECISE description of the FINISHED, BUILT house suitable for an AI image generator.

CRITICAL REQUIREMENT — ROOFLINE ACCURACY:
You MUST describe the roofline section by section from left to right. Many houses have MULTIPLE distinct roof planes at DIFFERENT heights — do NOT merge them into a single generic roofline. For example:
- A gable over the main body + a lower skillion or hip over the entry porch = TWO separate roof sections that must BOTH appear in your description
- A double gable = TWO gables at the same or different heights
- A hip with a lower skillion over the garage = TWO sections
If you see ANY step-down, break, or height change in the roofline, describe BOTH sections explicitly.

Describe the following precisely:
1. ROOFLINE: List each distinct roof section LEFT-TO-RIGHT. For each: shape (gable/hip/skillion/flat), relative height, and where it sits over the facade.
2. FACADE ZONES: garage position and width, entry porch location, main living section, any feature walls
3. WINDOWS: number, position, and approximate size of each window group
4. ENTRY: porch depth, columns/piers if any, front door position
5. PROPORTIONS: overall width vs height ratio

Target style: {style}
{f"Extra notes: {notes}" if notes else ""}
{f"Materials: {material_prompt}" if material_prompt else "Default materials: contemporary face brick walls, Colorbond steel sheet metal roof (NOT tiles), rendered feature panels."}
{zone_context}

Now write ONE coherent paragraph (4–6 sentences) for an AI image generator. Include ALL distinct roof sections. Do NOT mention the drawing, dimensions, annotations, or colour zones. Write as if describing a real photograph of the finished house."""


        vision_analysis = client.generate_text(vision_prompt, image_path=file_path)
        print(f"[Plans] Vision: {vision_analysis[:250]}...")

        # ── Step 2: Build generation prompt with explicit surface-zone mapping ──
        # Parse "zone: product_description; zone: product_description" format from frontend
        material_section_parts = []
        if material_prompt:
            for part in material_prompt.split(';'):
                part = part.strip()
                if ':' not in part:
                    continue
                zone, desc = part.split(':', 1)
                zone = zone.strip().lower()
                desc = desc.strip()
                # Map generic zone names to explicit surface descriptions
                zone_label = {
                    'brick':    'lower wall / primary facade surface',
                    'cladding': 'upper wall, feature panel, or secondary facade surface',
                    'roof':     'roof cladding (steel sheet, NOT roof tiles)',
                    'render':   'rendered or painted wall surface',
                }.get(zone, zone)
                material_section_parts.append(f"{zone_label}: {desc}")

        if material_section_parts:
            material_section = (
                " CRITICAL — render these exact materials on each surface zone: "
                + "; ".join(material_section_parts)
                + ". Reproduce the texture, profile, colour and finish of each product with photographic accuracy."
            )
        else:
            material_section = (
                " Contemporary face brick lower walls, smooth rendered upper feature panels, "
                "COLORBOND STEEL SHEET METAL ROOF (absolutely NOT roof tiles — it must be flat Colorbond sheeting), "
                "aluminium window frames, Colorbond fascia and gutters."
            )

        gen_prompt = (
            f"Photorealistic architectural exterior photograph of a newly built Australian residential house, "
            f"professional real estate photography, {style} style, "
            f"shot from the street, eye-level perspective, sunny day with blue sky and white clouds, "
            f"manicured front lawn, clean concrete driveway, established garden beds. "
            f"{vision_analysis} "
            f"IMPORTANT: reproduce the EXACT roofline geometry including ALL separate roof planes, breaks and step-downs — do NOT simplify or merge distinct roof sections. "
            f"{material_section} "
            f"Sharp focus, high detail, cinematic lighting, no text, no watermarks, no annotations, "
            f"photorealistic, 8K quality."
        )

        negative = (
            "drawing, sketch, blueprint, line art, architectural plan, annotations, dimensions, "
            "text labels, cartoon, render artefacts, low quality, blurry, watermark, signature, "
            "oversized trees, vegetation blocking facade, "
            "merged rooflines, incorrect roofline, simplified roof, wrong roof shape, "
            "missing roof section, extra roof section, tile roof, terracotta tiles"
        )

        # ── Step 3: Replicate ControlNet (best) or KIE text-to-image (fallback) ─
        replicate_key = os.environ.get("REPLICATE_API_TOKEN")

        if replicate_key:
            # PRIMARY: Flux Canny ControlNet — preserves elevation structure exactly
            # Upload to Supabase first so Replicate can fetch it via public URL
            # (Vercel serverless can't guarantee local file paths after request handling)
            try:
                with open(file_path, 'rb') as f_read:
                    file_bytes = f_read.read()
                unique_name = f"studio/elev_{uuid.uuid4().hex}.{ext}"
                if supabase:
                    supabase.storage.from_("cavalierone_uploads").upload(
                        unique_name, file_bytes, {"content-type": f"image/{ext}"}
                    )
                    control_image_url = supabase.storage.from_("cavalierone_uploads").get_public_url(unique_name)
                else:
                    control_image_url = request.host_url.rstrip('/') + f'/uploads/{safe_name}'

                engine_used = "replicate_controlnet"
                from utils.replicate_client import generate_image_controlnet
                image_url = generate_image_controlnet(
                    prompt=gen_prompt,
                    image_path=control_image_url,  # Pass public URL — works on Vercel
                    api_token=replicate_key
                )
            except Exception as rep_err:
                print(f"[Plans] Replicate failed ({rep_err}), falling back to KIE text-to-image")
                engine_used = "kie_text2img_fallback"
                from utils.kie_client import generate_image as kie_txt2img
                urls = kie_txt2img(
                    prompt=gen_prompt,
                    model='flux-2/flex-text-to-image',
                    aspect_ratio='16:9',
                    resolution='2K',
                    negative_prompt=negative,
                )
                image_url = urls[0] if urls else None

        else:
            # FALLBACK: KIE text-to-image from vision description
            # NOTE: We do NOT use img2img here — an elevation drawing on white paper
            # looks nothing like a real house photo, so img2img just colourises the drawing.
            # Text-to-image generates a proper render from the GPT-4 description.
            engine_used = "kie_text2img"
            from utils.kie_client import generate_image as kie_txt2img
            urls = kie_txt2img(
                prompt=gen_prompt,
                model='flux-2/flex-text-to-image',
                aspect_ratio='16:9',
                resolution='2K',
                negative_prompt=negative,
            )
            image_url = urls[0] if urls else None

        if not image_url:
            return jsonify({'error': 'AI generation returned no image'}), 500

        return jsonify({
            'image_url':       image_url,
            'vision_analysis': vision_analysis,
            'prompt':          gen_prompt,
            'engine':          engine_used,
        })

    except Exception as e:
        print(f"[Plans] Error: {e}")
        return jsonify({
            'error':  f'Generation failed: {str(e)}',
            'detail': traceback.format_exc()[-800:],
        }), 500


@app.route('/sales-training')
def sales_training():
    return render_template('sales_training.html')

@app.route('/image-studio')
def image_studio():
    """AI Image Studio — generate on-brand Cavalier marketing images via kie.ai."""
    if not user_can_access('rendering') and session.get('role') not in ('admin', 'marketing'):
        return "Access Denied: Image Studio requires admin or marketing access.", 403
    try:
        brand_config_path = os.path.join("resources", "brand_config.json")
        brand_config = {}
        if os.path.exists(brand_config_path):
            with open(brand_config_path, 'r') as f:
                brand_config = json.load(f)
    except Exception:
        brand_config = {}
    try:
        swatches_path = os.path.join("resources", "swatches.json")
        swatches = {}
        if os.path.exists(swatches_path):
            with open(swatches_path, 'r') as f:
                swatches = json.load(f)
    except Exception:
        swatches = {}
    return render_template(
        'image_generator.html',
        brand=brand_config,
        upload_limit_mb=UPLOAD_LIMIT_MB,
        swatches_json=json.dumps(swatches),
    )


@app.route('/sales-estimate')
def sales_estimate():
    if not user_can_access('sales_estimate'):
        return "Access Denied: Insufficient permissions for Sales Estimates.", 403
    """Sales Estimate Builder — interactive client-facing quote tool."""
    return render_template('sales_estimate.html')

@app.route('/sales-estimate/import', methods=['POST'])
def sales_estimate_import():
    """Parse an uploaded .xlsx sales estimate. Detects Display sheets, col B=desc, D=est, E=contract."""
    import io, traceback
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception as e:
        return jsonify({'error': f'Could not open file: {str(e)}'}), 400

    try:
        all_sheets = wb.sheetnames
        forced_sheet = request.form.get('sheet', '').strip()

        def find_est_con_cols(ws):
            """Scan first 15 rows for Estimate/Contract header; return (desc,est,con) indices or None."""
            for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
                vals = [str(c).strip().lower() if c is not None else '' for c in row]
                if 'estimate' in vals and 'contract' in vals:
                    ei = vals.index('estimate')
                    ci = vals.index('contract')
                    di = max(0, min(ei, ci) - 2)
                    return (di, ei, ci)
            return None

        def score_sheet(ws):
            """
            Score how likely this sheet is to be the actual client sales estimate.
            Key insight: Cavalier Homes workbooks contain many reference sheets
            (MALDON DISPLAY, PARKDALE DISPLAY, Lifestyle, Land) alongside the
            actual client estimate sheet. We want the client-specific one.
            """
            score = 0
            desc_col, est_col, con_col = 1, 3, 4
            found_header = False
            title_lower = ws.title.lower()

            # HARD PENALISE: known reference/pricing sheets — never the client estimate
            pricing_sheet_keywords = ['display', 'lifestyle', 'land', 'price list',
                                       'price guide', 'inclusions', 'options list']
            if any(k in title_lower for k in pricing_sheet_keywords):
                score -= 200  # basically disqualified unless nothing else exists

            # BOOST: sheet title contains client-sounding text (not a category name)
            generic_titles = {'sheet1','sheet2','sheet3','estimate','summary',
                              'quote','template','example','sample','master','data'}
            if title_lower not in generic_titles and len(ws.title) > 3:
                # If the sheet title looks like a name / job (not a category)
                if not any(k in title_lower for k in pricing_sheet_keywords):
                    score += 15  # likely named after the client or job

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
                vals = [str(c).strip() if c is not None else '' for c in row]
                vals_lower = [v.lower() for v in vals]
                row_str = ' '.join(vals_lower)

                # Strong signal: explicit Estimate + Contract column headers
                if 'estimate' in vals_lower and 'contract' in vals_lower:
                    ei = vals_lower.index('estimate')
                    ci = vals_lower.index('contract')
                    desc_col = max(0, min(ei, ci) - 2)
                    est_col  = ei
                    con_col  = ci
                    score += 50
                    found_header = True

                # Strong signal: actual client name row (not the Cavalier Homes office)
                if any(k in row_str for k in ('client/s name', 'client name', 'prepared for')):
                    score += 40

                # Strong signal: real job site address
                if 'site address' in row_str:
                    score += 30

                # Moderate signal: has numeric dollar values
                numeric_count = sum(1 for v in vals
                                    if v.replace(',','').replace('.','')
                                      .replace('-','').strip().isdigit())
                if numeric_count >= 2:
                    score += 3

                # Penalise: letter/intro content (not an estimate table)
                if any(k in row_str for k in ('dear ', 'yours sincerely', 'abn: ',
                                               'i am pleased', 'we are pleased')):
                    score -= 30

                # Penalise: base price / house model list (pricing catalogue, not estimate)
                if any(k in row_str for k in ('base price', 'lifestyle single storey',
                                               'house design', 'display pricing')):
                    score -= 60

            return score, desc_col, est_col, con_col

        # If user explicitly picked a sheet, use it directly
        if forced_sheet and forced_sheet in wb.sheetnames:
            ws = wb[forced_sheet]
            result = find_est_con_cols(ws)
            if result:
                desc_col, est_col, con_col = result
        else:
            # Score every sheet, pick the best
            best_score = -999
            ws = None
            desc_col, est_col, con_col = 1, 3, 4

            for sheet in wb.worksheets:
                s, dc, ec, cc = score_sheet(sheet)
                if s > best_score:
                    best_score = s
                    ws = sheet
                    desc_col, est_col, con_col = dc, ec, cc

            if ws is None:
                ws = wb.worksheets[0]

        def money(v):
            if v is None: return None
            s = str(v).strip()
            if s in ('', 'None', 'nan', '#REF!', '#VALUE!', 'POA', 'TBC'): return None
            try: return float(s.replace(',', '').replace('$', ''))
            except: return None

        def flag(raw, num):
            r = str(raw).strip().lower()
            if r in ('included', 'incl', 'incl.'): return 'Incl.'
            if r in ('tbc', 'tba'): return 'TBC'
            if r in ('allowance', 'allow', 'allow.', 'allowance only'): return 'Allow.'
            if r in ('note only', 'note'): return 'Note'
            return '$'

        def is_section(desc, est_raw):
            d = desc.strip()
            if not d or len(d) > 90: return False
            if d != d.upper(): return False
            noise = {'total','sub total','subtotal','grand total','gst','description','estimate','contract','difference','price'}
            if d.lower() in noise: return False
            if money(est_raw) is not None: return False
            return True

        # Rows to skip if description CONTAINS these (partial match)
        SKIP = ['progress payment','gst amount','percentage %','base stage','frame stage',
                'lock-up','fixing stage','completion stage','building contract price',
                'client/s signature','yours faithfully','thank you for','preliminary pricing',
                're: preliminary','dear ','i look forward','as discussed','once the initial',
                'if you are pleased','we will also','you wish to proceed',
                'all estimates are subject']

        # Rows to skip if description EXACTLY matches these (aggregate totals — already computed from line items)
        SKIP_EXACT = {'total','sub total','subtotal','grand total','sub-total',
                      'estimated total','contract total','estimate total',
                      'total estimate','total contract','includes gst',
                      'balance of contract'}

        rows_out, meta = [], {}

        for row in ws.iter_rows(values_only=True):
            row = list(row)
            max_idx = max(desc_col, est_col, con_col)
            while len(row) <= max_idx:
                row.append(None)

            d = str(row[desc_col]).strip() if row[desc_col] is not None else ''
            e = str(row[est_col]).strip()  if row[est_col]  is not None else ''
            c = str(row[con_col]).strip()  if row[con_col]  is not None else ''

            if not d and not e and not c: continue
            if any(p in d.lower() for p in SKIP): continue
            if d.lower() in SKIP_EXACT: continue
            if d.lower() in ('description','') and e.lower() in ('estimate','estimate $',''): continue

            # Extract meta fields (client name, site, date) — look in both directions
            dl = d.lower()
            if ('client/s name' in dl or 'client name' in dl or 'client:' in dl) and not meta.get('client'):
                # Value may be in estimate col position, or adjacent cells
                for ci2 in [est_col, desc_col+1, desc_col+2]:
                    if ci2 < len(row) and row[ci2] and str(row[ci2]).strip() not in ('', 'None'):
                        v = str(row[ci2]).strip()
                        if not any(x in v.lower() for x in ('date','address','$')):
                            meta['client'] = v; break
            if 'site address' in dl and not meta.get('site'):
                for ci2 in [est_col, desc_col+1, desc_col+2]:
                    if ci2 < len(row) and row[ci2] and str(row[ci2]).strip() not in ('', 'None'):
                        meta['site'] = str(row[ci2]).strip(); break
            if ('date' in dl or 'date:' in dl) and not meta.get('date'):
                for ci2 in [est_col, desc_col+1, desc_col+2]:
                    if ci2 < len(row) and row[ci2]:
                        v = str(row[ci2]).strip()
                        if v and v.lower() not in ('date','none',''):
                            meta['date'] = v; break
            if ('home design' in dl or 'house type' in dl or 'design:' in dl) and not meta.get('design'):
                for ci2 in [est_col, desc_col+1, desc_col+2]:
                    if ci2 < len(row) and row[ci2] and str(row[ci2]).strip() not in ('', 'None'):
                        meta['design'] = str(row[ci2]).strip(); break

            en = money(e)
            cn = money(c)

            # Skip meta label rows (no actual estimate value)
            if dl in ('client/s name:','client name:','site address:','date:','home design:','house type:'):
                continue

            if e.lower() in ('included','incl','incl.'): e_str = ''
            elif en is not None: e_str = f'{en:.2f}'
            elif e in ('#REF!','#VALUE!','nan','None','POA',''): e_str = ''
            else: e_str = e

            c_str = f'{cn:.2f}' if cn is not None else ('' if c in ('#REF!','#VALUE!','nan','None') else c)

            if is_section(d, e):
                rows_out.append({'t': 'sec', 'name': d})
            elif d:
                rows_out.append({'t':'item','desc':d,'flag':flag(e,en),'est':e_str,'con':c_str})

        return jsonify({
            'rows':       rows_out,
            'meta':       meta,
            'sheet':      ws.title,
            'total_rows': len(rows_out),
            'all_sheets': all_sheets,
        })

    except Exception as exc:
        return jsonify({'error': f'Parse error: {str(exc)}', 'detail': traceback.format_exc()[-600:]}), 500


@app.route('/create_manual', methods=['GET', 'POST'])
def create_manual():
    if request.method == 'GET':
        return render_template('manual_form.html')
    
    # POST
    title = request.form.get('title')
    subtitle = request.form.get('subtitle')
    raw_content = request.form.get('content')
    
    # Convert Markdown to HTML
    html_content = markdown.markdown(raw_content, extensions=['tables'])
    
    # Inject classes for styling (Hack to match Cavalier styles)
    # Replace <blockquote> with <div class="highlight-box">
    html_content = html_content.replace('<blockquote>', '<div class="highlight-box">').replace('</blockquote>', '</div>')
    
    return render_template('generic_manual_template.html', title=title, subtitle=subtitle, body_content=html_content)

@app.route('/generate_pdf', methods=['POST'])
def generate_pdf():
    # Extract data from hidden fields or re-collect
    data = request.form.to_dict()
    
    # Filename
    home_name = data.get('home_name', 'Brochure').replace(' ', '_').replace('(', '').replace(')', '')
    filename = f"{home_name}_Brochure.pdf"
    
    # Save to /tmp/ for serverless environments
    filepath = os.path.join(tempfile.gettempdir(), filename)
    
    # Render the HTML content for PDF
    # We reuse the logic from preview_brochure to prep data, or just extract here
    # Refactoring extraction to helper might be cleaner, but for now inline
    
    # --- Data Prep (Same as Preview) ---
    image_url = data.get('image_path') or ""
    plan_url = data.get('plan_path') or ""
    
    # Fallback Image Logic (if paths completely missing from form)
    if not image_url or not plan_url:
        # Re-run discovery if paths were lost (simulated re-discovery for statelessness)
        facade_dir = os.path.join(app.root_path, 'resources', 'sharepoint_files', 'Facade Renders')
        plans_dir = os.path.join(app.root_path, 'resources', 'sharepoint_files', 'Plans')
        home_name_clean = data.get('home_name', '').split(' ')[0]
        
        if not image_url and os.path.exists(facade_dir):
            for f in os.listdir(facade_dir):
                if home_name_clean.lower() in f.lower() and "The Rise" in f:
                    image_url = f"/resources/sharepoint_files/Facade Renders/{f}"
                    break
            if not image_url: # Fallback specific
                 image_url = "/resources/sharepoint_files/Facade Renders/Malvern - The Rise.png"
                 
        if not plan_url and os.path.exists(plans_dir):
             for f in os.listdir(plans_dir):
                if home_name_clean.lower() in f.lower() and f.lower().endswith(('.jpg', '.png')):
                     plan_url = f"/resources/sharepoint_files/Plans/{f}"
                     break

    # Context for Template
    context = {
        'title': data.get('home_name', 'Home Design'),
        'subtitle': data.get('lot_details', 'House & Land Package'),
        'hero_headline': 'New Home',
        'hero_image': image_url, # Pass URL. xhtml2pdf needs absolute paths usually, but we'll try relative first.
                                 # If xhtml2pdf fails with URLs, we might need absolute system paths.
        'bedrooms': data.get('bedrooms', 4),
        'bathrooms': data.get('bathrooms', 2),
        'car_spaces': data.get('cars', 2),
        'price': data.get('price_text', 'Call for Price'), # Add field for price
        'long_description': data.get('description', ''),
        'inclusions_summary': data.get('inclusions', '').split('\n'),
        'floorplan_image': plan_url,
        'contact_name': 'Sales Team',
        'contact_phone': '03 5823 1859',
        'contact_email': 'sales@cavalierhomes.com.au',
        'website_url': 'www.cavalierhomes.com.au',
        'legal_disclaimer_text': 'Images are for illustrative purposes. Refer to contract for full details.'
    }
    
    # Special Handler for xhtml2pdf images: It requires absolute SYSTEM paths for local files usually
    # We will convert the web urls (/resources/...) to system paths
    if context['hero_image'].startswith('/'):
        context['hero_image'] = os.path.join(app.root_path, context['hero_image'].lstrip('/'))
    if context['floorplan_image'].startswith('/'):
        context['floorplan_image'] = os.path.join(app.root_path, context['floorplan_image'].lstrip('/'))

    html_content = render_template('marketing_print.html', **context)
    
    try:
        from utils.pdf_generator import create_brochure_from_html
        success = create_brochure_from_html(filepath, html_content)
        
        if success:
            if supabase:
                # Upload PDF to Supabase
                unique_pdf_name = f"{uuid.uuid4().hex}_{filename}"
                with open(filepath, 'rb') as f:
                    supabase.storage.from_("cavalierone_pdfs").upload(unique_pdf_name, f.read(), {"content-type": "application/pdf"})
                pdf_url = supabase.storage.from_("cavalierone_pdfs").get_public_url(unique_pdf_name)
                # Cleanup local file
                try:
                    os.remove(filepath)
                except:
                    pass
                from flask import redirect
                return redirect(pdf_url)
            else:
                # Fallback to local send_file if no storage setup
                return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            return "Error generating PDF content", 500
    except Exception as e:
        return f"Error generating PDF: {str(e)}", 500

@app.route('/preview_brochure', methods=['POST'])
def preview_brochure():
    data = request.form.to_dict()
    home_name = data.get('home_name', 'Brochure')
    
    # --- 1. Facade Selection ---
    # Try to find a matching facade in 'Facade Renders'
    facade_dir = os.path.join(app.root_path, 'resources', 'sharepoint_files', 'Facade Renders')
    image_url = "" # Default empty
    
    # If image_name explicitly provided, use it
    # If image_name explicitly provided (legacy) or facade_select (new)
    if data.get('image_name'):
        image_url = f"/resources/sharepoint_files/Facade Renders/{data.get('image_name')}"
    elif data.get('facade_select'):
        # Construct filename from Home + Facade selection
        # Naming convention seems to be "HomeName - FacadeName.png" or "FacadeName Facade.png"
        # Based on file list: "Malvern - The Rise.png", "The Rise Facade.png"
        selected_facade = data.get('facade_select')
        
        # Try specific combo first: "Home Name - Facade.png"
        # Clean Home Name (remove size/brackets if needed, e.g. "Malvern 20" -> "Malvern")
        series_name = home_name.split(' ')[0] 
        combo_name_1 = f"{series_name} - {selected_facade}.png"
        combo_name_2 = f"{home_name} - {selected_facade}.png"
        
        # Try Generic Facade: "The Rise Facade.png"
        generic_name = f"{selected_facade} Facade.png"
        
        if os.path.exists(os.path.join(facade_dir, combo_name_1)):
             image_url = f"/resources/sharepoint_files/Facade Renders/{combo_name_1}"
        elif os.path.exists(os.path.join(facade_dir, combo_name_2)):
             image_url = f"/resources/sharepoint_files/Facade Renders/{combo_name_2}"
        elif os.path.exists(os.path.join(facade_dir, generic_name)):
             image_url = f"/resources/sharepoint_files/Facade Renders/{generic_name}"
        else:
             # Fallback to loose search
             image_url = "" # Will fall through to auto-discover
             print(f"Could not find exact match for facade: {selected_facade}")
    
    if not image_url:
        # Auto-discover
        design_series = home_name.split(' ')[0] 
        best_match = None
        
        if os.path.exists(facade_dir):
            for f in os.listdir(facade_dir):
                if design_series.lower() in f.lower():
                    # Prefer "The Rise" or "Standard"
                    if "The Rise" in f:
                        best_match = f
                        break # Perfect
                    if "Standard" in f:
                        best_match = f
                    if not best_match:
                        best_match = f # Take first loose match
        
        if best_match:
            image_url = f"/resources/sharepoint_files/Facade Renders/{best_match}"
        else:
             # Fallback
             if "malvern" in home_name.lower(): 
                 image_url = "/resources/sharepoint_files/Facade Renders/Malvern - The Rise.png"

    # --- 2. Floor Plan Selection ---
    plans_dir = os.path.join(app.root_path, 'resources', 'sharepoint_files', 'Plans')
    plan_url = ""
    
    if os.path.exists(plans_dir):
        # Strict check first
        for f in os.listdir(plans_dir):
            if home_name.lower() in f.lower() and f.lower().endswith(('.jpg', '.png', '.jpeg')):
                 plan_url = f"/resources/sharepoint_files/Plans/{f}"
                 break
        
        # If no strict match, try Series match
        if not plan_url:
             design_series = home_name.split(' ')[0]
             for f in os.listdir(plans_dir):
                if design_series.lower() in f.lower() and f.lower().endswith(('.jpg', '.png', '.jpeg')):
                     plan_url = f"/resources/sharepoint_files/Plans/{f}"
                     break

    # Context Mapping to Flat Schema
    context = {
        'title': home_name,
        'subtitle': data.get('lot_details', ''),
        'hero_image': image_url,
        'bedrooms': data.get('bedrooms', 4),
        'bathrooms': data.get('bathrooms', 2),
        'car_spaces': data.get('cars', 2),
        'price': 'Contact Agent', # Placeholder until form has price field
        'long_description': data.get('description', ''),
        'inclusions_summary': data.get('inclusions', '').split('\n') if data.get('inclusions') else [],
        'floorplan_image': plan_url,
        'contact_name': 'Sales Team',
        'contact_phone': '03 5823 1859',
        'contact_email': 'sales@cavalierhomes.com.au',
        'website_url': 'www.cavalierhomes.com.au',
        'legal_disclaimer_text': '*Images for illustrative purposes. Refer to contract.',
        
        # Pass raw paths back as hidden fields for PDF gen to reuse
        'image_path': image_url,
        'plan_path': plan_url,
        'home_name': home_name,
        'description': data.get('description'),
        'inclusions': data.get('inclusions'),
        'lot_details': data.get('lot_details')
    }

    return render_template('marketing_print.html', **context)

@app.route('/resources/<path:filename>')
def serve_resources(filename):
    resource_dir = os.path.join(app.root_path, "resources")
    return send_from_directory(resource_dir, filename)

@app.route('/generate_rendering', methods=['POST'])
def generate_rendering():
    from marketing.llm_client import LLMClient
    import requests
    import time
    
    # 1. Handle File Upload
    file = request.files.get('plan_image')
    style_preference = request.form.get('style', 'Modern Australian')
    additional_notes = request.form.get('notes', '')
    use_strict_mode = request.form.get('strict_mode') == 'true'
    
    if not file or not file.filename:
        return "Error: No file uploaded", 400
        
    # Save to resources/uploads
    upload_dir = os.path.join("resources", "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file.filename)
    file.save(file_path)
    
    # 2. Initialize Client
    client = LLMClient(provider="openai")
    
    # SPECIAL PATH: STRICT GEOMETRY MODE (Image Variation)
    if use_strict_mode:
        print("Using STRICT GEOMETRY MODE (Image Variation)")
        # DALL-E 2 Variation requires a Square PNG. We must convert/crop.
        # Simple center crop to square.
        from PIL import Image
        try:
            img = Image.open(file_path)
            
            # --- FIX: RESIZE BEFORE CROP ---
            # OpenAI requires < 4MB. 1024x1024 is safe.
            img.thumbnail((1024, 1024))
            
            # Resize/Crop to square
            min_dim = min(img.size)
            left = (img.width - min_dim)/2
            top = (img.height - min_dim)/2
            right = (img.width + min_dim)/2
            bottom = (img.height + min_dim)/2
            img_cropped = img.crop((left, top, right, bottom))
            
            # Save as PNG
            square_path = os.path.join(upload_dir, f"square_{file.filename}.png")
            img_cropped.save(square_path, format="PNG")
            
            # Generate Variation
            image_url = client.vary_image(square_path)
            
            # Use a dummy prompt for display since we didn't generate one
            rendering_prompt = f"Strict Geometry Mode: Variation of {file.filename}\n(Text Analysis Skipped)"
            
            if "Error" in image_url:
                 return f"Variation Failed: {image_url}", 500
                 
            # ... proceed to download (shared logic below) ...
            
        except Exception as e:
            with open("debug_log.txt", "a") as f:
                f.write(f"Strict Mode Error: {e}\n")
            return f"Strict Mode Error: {e}", 500
            
    else:
        # STANDARD PATH: Vision -> DALL-E 3
        print(f"Analyzing plan: {file_path}")
        
        # Step 1: Vision Analysis (Forensic Geometry Scanner)
        analysis_prompt = f"""
        Analyze this architectural image for a "Digital Twin" reconstruction. 
        Act as a Building Surveyor.
        
        Extract the following FORENSIC DETAILS:
        1. ROOF_FORM: Exact type (e.g. Hip, Gable). 
           - CRITICAL: Note the pitch/slope direction.
        2. EAVES:
           - Garage: Does it have eaves (overhang) or is it a "FLUSH EAVE"? 
             (Flush Eave = Roof ends exactly at external brick line with fascia/gutter).
           - Main House: Eave width.
        3. PIERS_AND_COLUMNS:
           - Default Assumption: FULL HEIGHT (Ground to Roof) unless explicitly low wall.
           - Material (Brick/Timber/Render)?
        4. POSITIVE_CONSTRUCTION: 
           - Describe the garage eave detail. "Flush Eave" or "Overhang"?
        5. MATERIALS: 
           - List materials.
           - IS THERE RENDER? (Yes/No). If NO, explicitly state "NO RENDER".
        6. STORY_COUNT: Single or Double Story?
        
        User Notes: {additional_notes}
        Target Style: {style_preference}
        
        Output concise, factual observations.
        """
        
        try:
            # Get the description from the Vision Model
            facade_analysis = client.generate_text(analysis_prompt, image_path=file_path)
            print(f"Vision Analysis: {facade_analysis}")
            
            # SAVE DEBUG LOG (Vital for refining the prompt)
            debug_path = os.path.join("resources", "debug_vision.txt")
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(f"--- Vision Analysis for {file.filename} ---\n")
                f.write(facade_analysis)
                f.write("\n-----------------------------------\n")
            
            # Step 2: Choose Engine (DALL-E 3 vs ControlNet)
            replicate_key = os.environ.get("REPLICATE_API_TOKEN")
            
            if replicate_key:
                print("--- SWITCHING TO CONTROLNET/REPLICATE MODE ---")
                from utils.replicate_client import generate_image_controlnet
                
                # For ControlNet, we want a simpler material prompt, not the structural list
                # because the lines ARE the structure.
                controlnet_prompt = f"Photorealistic 8k architectural render, {facade_analysis}. Best quality."
                
                image_url = generate_image_controlnet(
                    prompt=controlnet_prompt,
                    image_path=file_path,
                    api_token=replicate_key
                )
                rendering_prompt = controlnet_prompt # Set rendering_prompt for later use
            else:
                # Legacy DALL-E 3 Mode
                from utils.prompt_builder import build_image_prompt
                prompt_data = {
                    "home_name": "Custom Upload",
                    "facade_description": facade_analysis, 
                    "structural_details": facade_analysis,
                    "environment": "Manicured Australian suburban garden, sunny day"
                }
                
                rendering_prompt = build_image_prompt(prompt_data) # Use rendering_prompt here
                print(f"Generated Prompt: {rendering_prompt}")
                
                # Step 3: DALL-E 3 Generation
                # USE WIDE ASPECT RATIO (1792x1024) to prevent "second story hallucination" on wide houses
                print("Generating rendering...")
                # Enforce HD quality for production
                image_url = client.generate_image(rendering_prompt, size="1792x1024", quality="hd")

            print(f"Generated Image URL: {image_url}")
            
            # The original code had a return jsonify here, but the instruction implies
            # it should continue to the shared download logic.
            # I will keep the original flow for now, but if the user intended to return JSON,
            # they should clarify. Assuming the return jsonify was part of the instruction's
            # example of how the new logic *could* work, but not the final desired flow.
            # If it was meant to be a final return, the shared download logic would be unreachable.
            # For now, I'll integrate the logic and let it flow to the existing download.
            
        except Exception as e:
            with open("debug_log.txt", "a") as f:
                f.write(f"Standard Mode Error: {e}\n")
            print(f"Error generating rendering: {e}")
            import traceback
            traceback.print_exc()
            return f"Standard Mode Error: {e}", 500 # Keep original error return for consistency with shared logic

    # ... (Shared Download Logic) ...
    try:
        if "Error" in image_url or not image_url.startswith("http"):
            error_msg = str(image_url)
            if "billing_hard_limit_reached" in error_msg:
                 return render_template('output.html', prompt=rendering_prompt, content="<div style='color:red; background: #fee; padding:20px; border-radius:10px;'><h3>⚠️ OpenAI Billing Limit Reached</h3><p>Your OpenAI API account has run out of credits.</p><p>Please log in to <a href='https://platform.openai.com/settings/organization/billing' target='_blank'>platform.openai.com</a> and add credits to use DALL-E 3.</p></div>", data={'media_type': 'Error', 'home_name': 'Error'})
            if "content_policy_violation" in error_msg:
                 safe_prompt = rendering_prompt.replace('\n', '<br>')
                 return render_template('output.html', prompt=rendering_prompt, content=f"<div style='color:red; background: #fee; padding:20px; border-radius:10px;'><h3>⚠️ Safety Policy Violation</h3><p>The AI rejected the prompt as unsafe.</p><p><strong>Rejected Prompt:</strong><br><em>{safe_prompt}</em></p></div>", data={'media_type': 'Error', 'home_name': 'Error'})
    
            return f"Image Generation Failed: {image_url}", 500
            
        print(f"Generated Image URL: {image_url}")
        
        # 5. Download and Save Image Locally (Persistence)
        render_filename = f"render_{int(time.time())}.png"
        render_path = os.path.join(upload_dir, render_filename)
        
        # Ensure requests is imported
        import requests
        img_data = requests.get(image_url).content
        with open(render_path, 'wb') as handler:
            handler.write(img_data)
        
        print(f"Saved rendering to: {render_path}")
    
        # 6. Render Output
        user_data = {
            'media_type': 'Architectural Rendering',
            'tone': style_preference,
            'image': render_filename, 
            'notes': f"Source: {file.filename}"
        }
        
        formatted_analysis = rendering_prompt.replace('\n', '<br>')
        caption_content = f"<strong>AI Analysis & Prompt:</strong><br><br>{formatted_analysis}"
        
        return render_template('output.html', prompt=rendering_prompt, content=caption_content, data=user_data)
        
    except Exception as e:
        print(f"Error in rendering pipeline: {e}")
        return f"System Error: {str(e)}", 500

@app.route('/brain')
@require_role(['admin'])
def brain_dashboard():
    # 1. Fetch Notebooks
    notebooks = []
    try:
        notebooks = notebook_client.list_notebooks()
        # Handle case where it returns a dict with 'notebooks' key or similar, 
        # or if it returns an error dict.
        if isinstance(notebooks, dict):
            if 'notebooks' in notebooks:
                notebooks = notebooks['notebooks']
            elif 'error' in notebooks:
                print(f"Error fetching notebooks: {notebooks['error']}")
                notebooks = []
            elif 'content' in notebooks: 
                # MCP raw response might be complex, but notebook_client.list_notebooks 
                # usually calls _parse_content. Let's assume it returns the list or dict.
                pass
                
        # Ensure it's a list
        if not isinstance(notebooks, list):
            notebooks = []
            
    except Exception as e:
        print(f"Exception fetching notebooks: {e}")
        notebooks = []

    # 2. Fetch Recent Meetings (Fireflies)
    meetings = fireflies_client.get_recent_meetings(limit=5)
    
    return render_template('brain.html', meetings=meetings, notebooks=notebooks)

@app.route('/api/brain/create_notebook', methods=['POST'])
def api_create_notebook():
    title = request.form.get('title')
    resp = notebook_client.create_notebook(title)
    return jsonify(resp)

@app.route('/api/brain/add_source', methods=['POST'])
def api_add_source():
    notebook_id = request.form.get('notebook_id')
    url = request.form.get('url')
    text = request.form.get('text')
    title = request.form.get('title', 'Text Source')
    
    if url:
        resp = notebook_client.add_url_source(notebook_id, url)
    elif text:
        resp = notebook_client.add_text_source(notebook_id, text, title)
    else:
        return jsonify({'error': 'No source provided'}), 400
        
    return jsonify(resp)

@app.route('/api/brain/query', methods=['POST'])
def api_query_brain():
    notebook_id = request.form.get('notebook_id')
    question = request.form.get('question')
    resp = notebook_client.query(notebook_id, question)
    return jsonify(resp)

@app.route('/api/brain/import_meeting', methods=['POST'])
def api_import_meeting():
    notebook_id = request.form.get('notebook_id')
    transcript_id = request.form.get('transcript_id')
    
    # 1. Get Text
    text = fireflies_client.get_transcript_text(transcript_id)
    if not text:
        return jsonify({'error': 'Could not fetch transcript'}), 400
        
    # 2. Add to Notebook
    title = f"Meeting Transcript {transcript_id}"
    resp = notebook_client.add_text_source(notebook_id, text, title)
    
    return jsonify(resp)

@app.route('/api/brain/upload_file', methods=['POST'])
def api_upload_file():
    notebook_id = request.form.get('notebook_id')
    file = request.files.get('file')
    
    if not file:
        return jsonify({'error': 'No file provided'}), 400
        
    filename = file.filename
    ext = os.path.splitext(filename)[1].lower()
    
    # Save temporarily
    upload_dir = os.path.join("resources", "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, filename)
    file.save(file_path)
    
    try:
        text_content = ""
        
        if ext == '.pdf':
            # Extract text from PDF
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(file_path)
                full_text = []
                for page in doc:
                    full_text.append(page.get_text())
                text_content = "\n".join(full_text)
                doc.close()
            except ImportError:
                 # Fallback to PyPDF2 if fitz not available (but it is in requirements)
                 import PyPDF2
                 with open(file_path, 'rb') as f:
                     reader = PyPDF2.PdfReader(f)
                     full_text = []
                     for page in reader.pages:
                         full_text.append(page.extract_text())
                     text_content = "\n".join(full_text)
                     
        elif ext in ['.txt', '.md', '.csv', '.json', '.py', '.js', '.html']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                text_content = f.read()
        else:
             return jsonify({'error': 'Unsupported file type'}), 400
             
        if not text_content.strip():
             return jsonify({'error': 'Could not extract text from file'}), 400
             
        # Upload as Text Source
        # Note: NotebookLM has limits on text source size (~200k tokens?). 
        # Large files might need chunking, but for now we send as is.
        resp = notebook_client.add_text_source(notebook_id, text_content, filename)
        
        # Cleanup
        try:
            os.remove(file_path)
        except: pass
        
        return jsonify(resp)
        
    except Exception as e:
        print(f"Error processing file upload: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/brain/refresh_auth', methods=['POST'])
def api_refresh_auth():
    try:
        # 1. Run auth tool
        # We need to find the executable path similarly to how we found it before
        # Using environment var or default
        auth_cmd = "notebooklm-mcp-auth"
        # Try full path if known or just rely on PATH if added
        # Since we had issues finding it before, use the one from .env or hardcoded fallback
        mcp_cmd = Config.NOTEBOOKLM_MCP_COMMAND
        if mcp_cmd and "notebooklm-mcp.exe" in mcp_cmd:
            auth_cmd = mcp_cmd.replace("notebooklm-mcp.exe", "notebooklm-mcp-auth.exe")
            
        print(f"Running Auth Tool: {auth_cmd}")
        
        # Run it
        import subprocess
        # This might need to interact with a window, which is tricky from a service
        # But if user is local, it might just work.
        # However, `notebooklm-mcp-auth` usually opens browser or connects to debug port.
        # If it just connects to existing Chrome debug port, it should be fine.
        
        process = subprocess.run(auth_cmd, shell=True, capture_output=True, text=True)
        
        if process.returncode != 0:
            return jsonify({'error': f"Auth tool failed: {process.stderr}"}), 500
            
        # 2. Copy file (Crucial Step: Sync .notebooklm-mcp/auth.json -> .notebooklm-auth.json)
        # We know where it saves by default now: ~/.notebooklm-mcp/auth.json
        # And where server looks: ~/.notebooklm-auth.json
        
        import shutil
        user_home = os.path.expanduser("~")
        src = os.path.join(user_home, ".notebooklm-mcp", "auth.json")
        dst = os.path.join(user_home, ".notebooklm-auth.json")
        
        if os.path.exists(src):
            shutil.copy2(src, dst)
            return jsonify({'status': 'success', 'message': 'Auth refreshed and synced.'})
        else:
            return jsonify({'error': 'Auth file not found in new location.'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/render_document', methods=['POST'])
def render_document():
    """Accepts edited markdown from the draft editor and renders the final document."""
    title = request.form.get('title', 'Generated Document')
    subtitle = request.form.get('subtitle', '')
    markdown_content = request.form.get('markdown_content', '')

    # Normalise non-standard bullet characters so the markdown parser
    # produces proper <ul><li> elements rather than a single paragraph.
    import re
    # Replace any line starting with ▸ (with optional leading whitespace) → '- '
    markdown_content = re.sub(r'^([ \t]*)▸\s*', r'\1- ', markdown_content, flags=re.MULTILINE)
    # Also catch • used as a bullet (outside standard markdown)
    markdown_content = re.sub(r'^([ \t]*)•\s*', r'\1- ', markdown_content, flags=re.MULTILINE)

    html_content = markdown.markdown(markdown_content, extensions=['tables'])

    # Inject highlight-box styling for blockquotes
    html_content = html_content.replace('<blockquote>', '<div class="highlight-box">').replace('</blockquote>', '</div>')

    return render_template('generic_manual_template.html', title=title, subtitle=subtitle, body_content=html_content)

@app.route('/api/doc/get_meetings')
def api_doc_get_meetings():
    """Returns recent Fireflies meetings as JSON for the Doc Formatter picker."""
    try:
        meetings = fireflies_client.get_recent_meetings(limit=20)
        return jsonify(meetings)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/doc/fetch_transcript')
def api_doc_fetch_transcript():
    """Fetches a Fireflies transcript by ID and returns its full text."""
    transcript_id = request.args.get('id')
    if not transcript_id:
        return jsonify({'error': 'No transcript ID provided'}), 400
    try:
        text = fireflies_client.get_transcript_text(transcript_id)
        if not text:
            return jsonify({'error': 'Could not fetch transcript — it may still be processing.'}), 404
        return jsonify({'text': text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Always resolve relative to this file so it works regardless of CWD
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
COMPANY_INFO_PATH = os.path.join(BASE_DIR, 'resources', 'company_info.md')

# ---------------------------------------------------------------------------
# STAFF LIST — Supabase-backed (Vercel filesystem is read-only at runtime)
# The company_staff table is the source of truth for all staff operations.
# company_info.md is only used as a read-only fallback if Supabase is down.
# ---------------------------------------------------------------------------

def get_staff_from_supabase():
    """
    Returns staff list from the Supabase company_staff table.
    Returns a list of dicts: [{name, role, category}]
    """
    if not supabase:
        return None  # Supabase not configured — caller will fall back to MD
    try:
        res = supabase.table('company_staff').select('name, role, category').order('category').order('name').execute()
        return res.data if res.data else []
    except Exception as e:
        print(f"[staff] Supabase read error: {e}")
        return None


def parse_staff_from_md():
    """
    Fallback: Parses the Team Structure section of company_info.md.
    Only used when Supabase is unavailable.
    Returns a list of dicts: [{name, role, category}]
    """
    import re
    staff = []
    if not os.path.exists(COMPANY_INFO_PATH):
        return staff

    with open(COMPANY_INFO_PATH, 'r', encoding='utf-8') as f:
        content = f.read()

    team_match = re.search(r'## Team Structure(.*?)(?=^## |\Z)', content, re.DOTALL | re.MULTILINE)
    if not team_match:
        return staff

    section = team_match.group(1)
    current_category = 'Staff'
    name_first_categories = {'Directors', 'Additional Staff', 'Sales Team'}

    for line in section.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        heading = re.match(r'^\*\*(.+?):\*\*\s*$', stripped)
        if heading:
            current_category = heading.group(1).strip()
            continue
        bullet = re.match(r'^[-*]\s+\*\*(.+?):\*\*\s*(.+)', stripped)
        if bullet:
            bold_text = bullet.group(1).strip()
            value_text = bullet.group(2).strip().rstrip('.')
            if current_category in name_first_categories:
                name, role = bold_text, value_text
            else:
                role, name = bold_text, value_text
            if name and role:
                staff.append({'name': name, 'role': role, 'category': current_category})

    return staff


def get_all_staff():
    """
    Primary entrypoint for reading the staff list.
    Uses Supabase if available, falls back to company_info.md.
    """
    staff = get_staff_from_supabase()
    if staff is not None:
        return staff
    # Supabase unavailable — use local file (will work locally, not on Vercel)
    return parse_staff_from_md()


@app.route('/api/staff')
def api_get_staff():
    """Returns the full staff list (Supabase → company_info.md fallback)."""
    try:
        staff = get_all_staff()
        return jsonify(staff)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/staff/add', methods=['POST'])
def api_add_staff():
    """
    Adds a new staff member to the company_staff Supabase table.
    Body: { name, role }
    """
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON body provided'}), 400

    name = (data.get('name') or '').strip()
    role = (data.get('role') or '').strip()

    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if not role:
        role = 'Staff Member'

    if not supabase:
        return jsonify({'error': 'Database not configured — cannot save new staff on this deployment.'}), 500

    try:
        # Check if already exists (case-insensitive)
        existing_res = supabase.table('company_staff').select('name').ilike('name', name).execute()
        if existing_res.data:
            all_staff = get_all_staff()
            return jsonify({'already_exists': True, 'staff': all_staff})

        # Insert new staff member
        supabase.table('company_staff').insert({
            'name': name,
            'role': role,
            'category': 'Additional Staff'
        }).execute()

        updated = get_all_staff()
        return jsonify({'added': True, 'staff': updated})

    except Exception as e:
        print(f"[staff/add] Error: {e}")
        return jsonify({'error': f'Could not save staff member: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)

